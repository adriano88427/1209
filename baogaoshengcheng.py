"""
优化后的报告生成模块
支持多种报告格式和自定义模板
"""

import pandas as pd
import numpy as np
import os
import time
import logging
from abc import ABC, abstractmethod
from typing import Dict, List, Any, Optional, Union, Callable
from pathlib import Path
from datetime import datetime
import json
import yaml
from jinja2 import Environment, FileSystemLoader, Template
import base64
import io

# 导入基础系统组件
from jichuxitong import ConfigManager, CacheManager, ErrorHandler, FactorAnalysisError, DataProcessingError, exception_handler, cache_result
import multiprocessing as mp
from functools import partial
import hashlib
import pickle
import json
import yaml
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
import jinja2
import base64
from io import BytesIO

# 导入系统组件
from enhanced_factor_analysis_system import (
    ConfigManager, CacheManager, ErrorHandler, FactorAnalysisError,
    AnalysisError, exception_handler, cache_result
)

class ReportFormat(ABC):
    """抽象报告格式"""
    
    @abstractmethod
    def generate(self, data: Dict[str, Any], template_path: str = None) -> str:
        """生成报告"""
        pass
    
    @abstractmethod
    def get_extension(self) -> str:
        """获取文件扩展名"""
        pass

class HTMLReportFormat(ReportFormat):
    """HTML报告格式"""
    
    def __init__(self, template_dir: str = "templates"):
        self.template_dir = template_dir
        self.env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(template_dir),
            autoescape=jinja2.select_autoescape(['html', 'xml'])
        )
    
    def generate(self, data: Dict[str, Any], template_path: str = None) -> str:
        """生成HTML报告"""
        if template_path is None:
            template_path = "report_template.html"
        
        try:
            template = self.env.get_template(template_path)
            html_content = template.render(**data)
            return html_content
        except Exception as e:
            logging.error(f"生成HTML报告失败: {e}")
            raise AnalysisError(f"生成HTML报告失败: {e}")
    
    def get_extension(self) -> str:
        """获取文件扩展名"""
        return ".html"

class PDFReportFormat(ReportFormat):
    """PDF报告格式"""
    
    def __init__(self, html_format: HTMLReportFormat):
        self.html_format = html_format
    
    def generate(self, data: Dict[str, Any], template_path: str = None) -> bytes:
        """生成PDF报告"""
        try:
            # 先生成HTML
            html_content = self.html_format.generate(data, template_path)
            
            # 使用weasyprint将HTML转换为PDF
            try:
                import weasyprint
                pdf_content = weasyprint.HTML(string=html_content).write_pdf()
                return pdf_content
            except ImportError:
                logging.warning("未安装weasyprint，无法生成PDF，返回HTML内容")
                return html_content.encode('utf-8')
        except Exception as e:
            logging.error(f"生成PDF报告失败: {e}")
            raise AnalysisError(f"生成PDF报告失败: {e}")
    
    def get_extension(self) -> str:
        """获取文件扩展名"""
        return ".pdf"

class ExcelReportFormat(ReportFormat):
    """Excel报告格式"""
    
    def __init__(self):
        pass
    
    def generate(self, data: Dict[str, Any], template_path: str = None) -> bytes:
        """生成Excel报告"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 为每个数据集创建工作表
            for sheet_name, sheet_data in data.items():
                if isinstance(sheet_data, pd.DataFrame):
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 添加DataFrame数据
                    for r in dataframe_to_rows(sheet_data, index=True, header=True):
                        ws.append(r)
                elif isinstance(sheet_data, dict):
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 添加字典数据
                    for key, value in sheet_data.items():
                        ws.append([key, str(value)])
                elif isinstance(sheet_data, list):
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 添加列表数据
                    for item in sheet_data:
                        if isinstance(item, dict):
                            # 如果是字典，添加键值对
                            for key, value in item.items():
                                ws.append([key, str(value)])
                            ws.append([])  # 添加空行分隔
                        else:
                            ws.append([str(item)])
            
            # 保存到字节流
            excel_stream = BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            
            return excel_stream.read()
        except Exception as e:
            logging.error(f"生成Excel报告失败: {e}")
            raise AnalysisError(f"生成Excel报告失败: {e}")
    
    def get_extension(self) -> str:
        """获取文件扩展名"""
        return ".xlsx"

class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, config_manager: ConfigManager, cache_manager: CacheManager, 
                 error_handler: ErrorHandler):
        self.config = config_manager
        self.cache = cache_manager
        self.error_handler = error_handler
        
        # 初始化报告格式
        self.formats = {}
        self._initialize_formats()
        
        # 注册配置监听器
        self.config.watch('reports.default_format', self._update_default_format)
    
    def _initialize_formats(self):
        """初始化报告格式"""
        # HTML格式
        template_dir = self.config.get('reports.template_dir', 'templates')
        html_format = HTMLReportFormat(template_dir)
        self.formats['html'] = html_format
        
        # PDF格式
        pdf_format = PDFReportFormat(html_format)
        self.formats['pdf'] = pdf_format
        
        # Excel格式
        excel_format = ExcelReportFormat()
        self.formats['excel'] = excel_format
    
    def _update_default_format(self, key: str, new_value: Any, old_value: Any):
        """更新默认报告格式"""
        logging.info(f"更新默认报告格式: {old_value} -> {new_value}")
    
    @exception_handler()
    def generate_factor_analysis_report(self, factor_data: Dict[str, pd.DataFrame], 
                                       return_data: pd.DataFrame, 
                                       factor_scores: Dict[str, pd.Series],
                                       factor_rankings: pd.DataFrame,
                                       format: str = None,
                                       output_path: str = None) -> str:
        """生成因子分析报告"""
        if format is None:
            format = self.config.get('reports.default_format', 'html')
        
        if format not in self.formats:
            raise AnalysisError(f"不支持的报告格式: {format}")
        
        # 尝试从缓存加载
        cache_key = f"factor_analysis_report_{format}_{hash(str(factor_data))}_{hash(str(return_data))}"
        cached_result = self.cache.get(cache_key)
        if cached_result is not None:
            logging.info(f"从缓存加载{format}格式的因子分析报告")
            return cached_result
        
        # 准备报告数据
        report_data = self._prepare_factor_analysis_report_data(
            factor_data, return_data, factor_scores, factor_rankings
        )
        
        # 生成报告
        try:
            report_content = self.formats[format].generate(report_data)
            
            # 缓存结果
            self.cache.set(cache_key, report_content)
            
            # 保存到文件
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"factor_analysis_report_{timestamp}{self.formats[format].get_extension()}"
            
            self._save_report(report_content, output_path, format)
            
            logging.info(f"生成{format}格式的因子分析报告: {output_path}")
            return output_path
        except Exception as e:
            logging.error(f"生成因子分析报告失败: {e}")
            self.error_handler.handle_error(e, "生成因子分析报告失败")
            raise
    
    def _prepare_factor_analysis_report_data(self, factor_data: Dict[str, pd.DataFrame], 
                                            return_data: pd.DataFrame,
                                            factor_scores: Dict[str, pd.Series],
                                            factor_rankings: pd.DataFrame) -> Dict[str, Any]:
        """准备因子分析报告数据"""
        # 基本统计信息
        basic_stats = {}
        for factor_name, data in factor_data.items():
            basic_stats[factor_name] = {
                '数据量': len(data),
                '股票数量': len(data.columns),
                '开始日期': data.index[0].strftime('%Y-%m-%d'),
                '结束日期': data.index[-1].strftime('%Y-%m-%d'),
                '缺失值': data.isnull().sum().sum(),
                '均值': data.mean().mean(),
                '标准差': data.std().mean(),
                '最小值': data.min().min(),
                '最大值': data.max().max()
            }
        
        # 因子评分统计
        score_stats = {}
        for factor_name, scores in factor_scores.items():
            score_stats[factor_name] = {
                '均值': scores.mean(),
                '标准差': scores.std(),
                '最小值': scores.min(),
                '最大值': scores.max(),
                '偏度': scores.skew(),
                '峰度': scores.kurtosis()
            }
        
        # 因子排名
        top_factors = factor_rankings.head(10)
        
        # 图表数据
        charts = self._generate_charts(factor_data, return_data, factor_scores)
        
        # 报告元数据
        metadata = {
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '因子数量': len(factor_data),
            '分析周期': f"{factor_data[list(factor_data.keys())[0]].index[0].strftime('%Y-%m-%d')} 至 {factor_data[list(factor_data.keys())[0]].index[-1].strftime('%Y-%m-%d')}"
        }
        
        return {
            'title': '因子分析报告',
            'metadata': metadata,
            'basic_stats': basic_stats,
            'score_stats': score_stats,
            'top_factors': top_factors.to_dict(),
            'charts': charts
        }
    
    def _generate_charts(self, factor_data: Dict[str, pd.DataFrame], 
                        return_data: pd.DataFrame,
                        factor_scores: Dict[str, pd.Series]) -> Dict[str, str]:
        """生成图表"""
        charts = {}
        
        # 设置图表样式
        plt.style.use('seaborn')
        
        # 因子分布图
        for factor_name, data in factor_data.items():
            if len(factor_data) <= 3:  # 限制图表数量
                plt.figure(figsize=(10, 6))
                
                # 随机选择几只股票展示
                sample_stocks = np.random.choice(data.columns, min(5, len(data.columns)), replace=False)
                
                for stock in sample_stocks:
                    plt.plot(data.index, data[stock], label=stock, alpha=0.7)
                
                plt.title(f'{factor_name} 因子值分布')
                plt.xlabel('日期')
                plt.ylabel('因子值')
                plt.legend()
                plt.grid(True)
                
                # 保存为base64字符串
                buffer = BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                buffer.seek(0)
                chart_data = base64.b64encode(buffer.read()).decode('utf-8')
                charts[f'{factor_name}_distribution'] = f"data:image/png;base64,{chart_data}"
                
                plt.close()
        
        # 因子评分分布图
        for factor_name, scores in factor_scores.items():
            if len(factor_scores) <= 3:  # 限制图表数量
                plt.figure(figsize=(10, 6))
                plt.hist(scores.dropna(), bins=30, alpha=0.7, edgecolor='black')
                plt.title(f'{factor_name} 评分分布')
                plt.xlabel('评分')
                plt.ylabel('频数')
                plt.grid(True)
                
                # 保存为base64字符串
                buffer = BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                buffer.seek(0)
                chart_data = base64.b64encode(buffer.read()).decode('utf-8')
                charts[f'{factor_name}_score_distribution'] = f"data:image/png;base64,{chart_data}"
                
                plt.close()
        
        return charts
    
    def _save_report(self, content: Union[str, bytes], output_path: str, format: str):
        """保存报告"""
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            if format == 'excel':
                with open(output_path, 'wb') as f:
                    f.write(content)
            else:
                with open(output_path, 'w', encoding='utf-8') as f:
                    if isinstance(content, bytes):
                        f.write(content.decode('utf-8'))
                    else:
                        f.write(content)
        except Exception as e:
            logging.error(f"保存报告失败: {e}")
            raise AnalysisError(f"保存报告失败: {e}")
    
    @exception_handler()
    def generate_performance_report(self, performance_data: Dict[str, pd.DataFrame],
                                  format: str = None,
                                  output_path: str = None) -> str:
        """生成性能报告"""
        if format is None:
            format = self.config.get('reports.default_format', 'html')
        
        if format not in self.formats:
            raise AnalysisError(f"不支持的报告格式: {format}")
        
        # 准备报告数据
        report_data = self._prepare_performance_report_data(performance_data)
        
        # 生成报告
        try:
            report_content = self.formats[format].generate(report_data)
            
            # 保存到文件
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"performance_report_{timestamp}{self.formats[format].get_extension()}"
            
            self._save_report(report_content, output_path, format)
            
            logging.info(f"生成{format}格式的性能报告: {output_path}")
            return output_path
        except Exception as e:
            logging.error(f"生成性能报告失败: {e}")
            self.error_handler.handle_error(e, "生成性能报告失败")
            raise
    
    def _prepare_performance_report_data(self, performance_data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """准备性能报告数据"""
        # 报告元数据
        metadata = {
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '性能指标数量': len(performance_data)
        }
        
        return {
            'title': '性能报告',
            'metadata': metadata,
            'performance_data': performance_data
        }
    
    def add_format(self, name: str, format: ReportFormat):
        """添加报告格式"""
        self.formats[name] = format
        logging.info(f"添加报告格式: {name}")
    
    def remove_format(self, name: str):
        """移除报告格式"""
        if name in self.formats:
            del self.formats[name]
            logging.info(f"移除报告格式: {name}")
    
    def get_format_names(self) -> List[str]:
        """获取报告格式名称列表"""
        return list(self.formats.keys())

# 示例使用
if __name__ == "__main__":
    # 设置日志
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # 初始化组件
    config_manager = ConfigManager("config.yaml", "production")
    cache_manager = CacheManager("cache", memory_size=50, disk_size=500)
    error_handler = ErrorHandler(config_manager)
    
    # 创建报告生成器
    report_generator = ReportGenerator(config_manager, cache_manager, error_handler)
    
    # 模拟数据
    dates = pd.date_range('2020-01-01', '2021-12-31', freq='D')
    stocks = [f'stock_{i}' for i in range(100)]
    
    # 创建因子数据
    factor_data = {
        'factor_1': pd.DataFrame(
            np.random.randn(len(dates), len(stocks)),
            index=dates,
            columns=stocks
        ),
        'factor_2': pd.DataFrame(
            np.random.randn(len(dates), len(stocks)) * 0.5,
            index=dates,
            columns=stocks
        )
    }
    
    # 创建收益数据
    return_data = pd.DataFrame(
        np.random.randn(len(dates), len(stocks)) * 0.01,
        index=dates,
        columns=stocks
    )
    
    # 创建因子评分数据
    factor_scores = {
        'factor_1': pd.Series(np.random.randn(len(stocks)), index=stocks),
        'factor_2': pd.Series(np.random.randn(len(stocks)), index=stocks)
    }
    
    # 创建因子排名数据
    factor_rankings = pd.DataFrame({
        'stock': stocks,
        'factor_1': np.random.randint(1, 100, len(stocks)),
        'factor_2': np.random.randint(1, 100, len(stocks))
    }).set_index('stock')
    
    # 生成报告
    try:
        # 生成HTML格式的因子分析报告
        html_report = report_generator.generate_factor_analysis_report(
            factor_data, return_data, factor_scores, factor_rankings, format='html'
        )
        print(f"生成HTML报告: {html_report}")
        
        # 生成Excel格式的因子分析报告
        excel_report = report_generator.generate_factor_analysis_report(
            factor_data, return_data, factor_scores, factor_rankings, format='excel'
        )
        print(f"生成Excel报告: {excel_report}")
        
    except Exception as e:
        print(f"生成报告失败: {e}")