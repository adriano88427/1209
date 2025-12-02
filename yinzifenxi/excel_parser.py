# -*- coding: utf-8 -*-
"""
复杂 Excel 表格解析器

该模块集中处理 Excel/CSV 数据的加载、列名归一化、类型转换以及解析诊断。
主分析流程只需要调用 load_excel_sources 即可获取统一格式的数据。
"""
from __future__ import annotations

import os
import unicodedata
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple, Any

import numpy as np
import pandas as pd
from pandas.api.types import infer_dtype

try:
    import pyarrow.xlsx  # type: ignore
    _HAS_PYARROW_XLSX = True
except Exception:  # pragma: no cover
    _HAS_PYARROW_XLSX = False

from openpyxl import load_workbook

DEFAULT_PARSE_CONFIG: Dict[str, Any] = {
    "preferred_engines": ["pyarrow", "openpyxl", "pandas"],
    "sheet_policy": "all",  # 'all' / 'first' / 'largest' / ['Sheet1', ...]
    "na_values": ["", " ", "NA", "N/A", "-", "—", "--"],
    "column_aliases": {},
    "column_types": {},
    "unit_rules": {"万": 1e4, "億": 1e8, "亿": 1e8},
    "column_parsers": {},
    "strict_mode": False,
    "percent_threshold": 0.8,
}


@dataclass
class ParseDiagnostics:
    file_path: str
    sheet_name: str
    rows: int
    columns: int
    present_columns: List[str] = field(default_factory=list)
    missing_fields: List[str] = field(default_factory=list)
    unmapped_columns: List[str] = field(default_factory=list)
    conversion_failures: Dict[str, float] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)


@dataclass
class ParsedData:
    data: pd.DataFrame
    diagnostics: List[ParseDiagnostics]
    unavailable_columns: List[str]


@dataclass
class NormalizationInfo:
    column: str
    semantic: str = "numeric"
    display: str = "raw"
    applied_scale: Optional[float] = None
    detected_percent_pattern: bool = False
    suspected_shrink: bool = False
    notes: Optional[str] = None


class FactorNormalizer:
    """简单的列正则化器，用于处理百分比/涨跌幅等列."""

    def __init__(self, percent_keywords: Optional[Sequence[str]] = None):
        if percent_keywords is None:
            percent_keywords = [
                "%",
                "比例",
                "占比",
                "收益",
                "涨幅",
                "跌幅",
                "回撤",
            ]
        self.percent_keywords = tuple(percent_keywords)

    def normalize(self, column: str, series: pd.Series, diagnostics=None) -> Tuple[pd.Series, NormalizationInfo]:
        info = NormalizationInfo(column=column)
        if series is None:
            return pd.Series(dtype=float), info

        ser = series.copy()
        if not isinstance(ser, pd.Series):
            ser = pd.Series(ser)

        percent_mask = pd.Series(False, index=ser.index)
        has_percent_symbol = False
        if ser.dtype == object:
            str_series = ser.astype(str).str.strip()
            percent_mask = str_series.str.contains("%", na=False)
            has_percent_symbol = percent_mask.any()
            cleaned = str_series.str.replace("%", "", regex=False).str.replace(",", "")
            ser_numeric = pd.to_numeric(cleaned, errors="coerce")
        else:
            ser_numeric = pd.to_numeric(ser, errors="coerce")
            percent_mask = pd.Series(False, index=ser.index)

        scale = 1.0
        if has_percent_symbol:
            ser_numeric.loc[percent_mask] = ser_numeric.loc[percent_mask] / 100.0
            scale *= 0.01
            info.detected_percent_pattern = True
            info.semantic = "percent"
            info.display = "ratio"

        if info.semantic == "numeric":
            info.display = "raw"

        if self._looks_like_percent_column(column):
            info.semantic = "percent"
            info.display = "ratio"
            abs_max = float(ser_numeric.abs().max(skipna=True)) if ser_numeric.notna().any() else 0.0
            if abs_max > 1.5:
                ser_numeric = ser_numeric / 100.0
                scale *= 0.01
            elif abs_max > 0 and abs_max < 0.5 and has_percent_symbol:
                # already scaled, do nothing
                pass

        abs_max_final = float(ser_numeric.abs().max(skipna=True)) if ser_numeric.notna().any() else 0.0
        if info.semantic == "percent" and abs_max_final > 0 and abs_max_final <= 0.01:
            info.suspected_shrink = True
            warning = f"列 {column} 最大值仅 {abs_max_final*100:.2f}%，疑似被错误除以 100"
            info.notes = warning if not info.notes else f"{info.notes}; {warning}"
            print(f"[WARN] {warning}")

        if scale != 1.0 or has_percent_symbol:
            info.applied_scale = scale

        return ser_numeric, info

    def _looks_like_percent_column(self, column: str) -> bool:
        if not column:
            return False
        return any(keyword in column for keyword in self.percent_keywords)

    def format_range(
        self,
        column: str,
        min_val: float,
        max_val: float,
        semantic: Optional[str] = None,
        decimals: int = 3,
    ) -> str:
        try:
            low = float(min_val)
            high = float(max_val)
        except (TypeError, ValueError):
            return "N/A"
        if not np.isfinite(low) or not np.isfinite(high):
            return "N/A"

        use_percent = False
        if semantic == "percent":
            use_percent = True
        elif self._looks_like_percent_column(column):
            use_percent = True

        if use_percent:
            return f"[{low * 100:.2f}%, {high * 100:.2f}%]"
        return f"[{low:.{decimals}f}, {high:.{decimals}f}]"


class MultiEngineExcelLoader:
    """负责根据配置自动选择引擎读取 Excel / CSV 文件。"""

    def __init__(self, config: Dict[str, Any]):
        self.config = config

    def read_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(file_path)

        ext = os.path.splitext(file_path)[1].lower()
        if ext in [".csv", ".txt"]:
            df = pd.read_csv(
                file_path,
                na_values=self.config.get("na_values"),
                keep_default_na=True,
                dtype=str,
            )
            return {"CSV": df}

        loaders = self.config.get("preferred_engines", [])
        for engine in loaders:
            try:
                if engine == "pyarrow" and _HAS_PYARROW_XLSX:
                    tables = pyarrow.xlsx.read_excel(file_path)  # type: ignore[attr-defined]
                    result = {}
                    for sheet_name, table in tables.items():
                        result[sheet_name] = table.to_pandas()
                    return result
                if engine == "openpyxl":
                    wb = load_workbook(file_path, data_only=True)
                    return {
                        name: pd.DataFrame(wb[name].values)
                        for name in wb.sheetnames
                    }
                if engine == "pandas":
                    sheets = pd.read_excel(
                        file_path,
                        sheet_name=None,
                        dtype=str,
                        na_values=self.config.get("na_values"),
                        keep_default_na=True,
                    )
                    return sheets
            except Exception:
                continue

        raise RuntimeError(f"无法使用任何引擎解析文件: {file_path}")


class ColumnNormalizer:
    """对列名执行 Unicode 规整、别名映射和模糊匹配。"""

    def __init__(self, column_aliases: Dict[str, Sequence[str]]):
        self.alias_map = self._build_alias_lookup(column_aliases)
        self.standard_keys = set(self.alias_map.keys())

    @staticmethod
    def _normalize_key(name: str) -> str:
        if name is None:
            return ""
        name = str(name)
        name = unicodedata.normalize("NFKC", name)
        name = name.replace("\n", " ").replace("\t", " ")
        return " ".join(name.split())

    def _build_alias_lookup(self, column_aliases: Dict[str, Sequence[str]]) -> Dict[str, str]:
        lookup = {}
        for standard, aliases in column_aliases.items():
            lookup[self._normalize_key(standard)] = standard
            for alias in aliases:
                lookup[self._normalize_key(alias)] = standard
        return lookup

    def map_columns(self, columns: Sequence[str]) -> Tuple[List[str], List[str]]:
        mapped = []
        unmapped = []
        for col in columns:
            norm = self._normalize_key(col)
            standard = self.alias_map.get(norm, norm)
            if standard == "":
                unmapped.append(str(col))
            mapped.append(standard if standard else str(col))
        return mapped, unmapped


class ValueConverter:
    """负责根据列类型将字符串转换为数值/日期等。"""

    def __init__(self, config: Dict[str, Any]):
        self.column_types = config.get("column_types", {})
        self.unit_rules = config.get("unit_rules", {})
        self.custom_parsers = config.get("column_parsers", {})
        self.na_values = set(config.get("na_values", []))

    def convert(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, float]]:
        df = df.copy()
        fail_rates: Dict[str, float] = {}

        for column in df.columns:
            series = df[column]
            if column in self.custom_parsers:
                parser = self.custom_parsers[column]
                try:
                    df[column] = parser(series)
                except Exception:
                    fail_rates[column] = 1.0
                continue

            target_type = self.column_types.get(column, "auto")
            converted, fail_rate = self._convert_series(series, target_type, column)
            df[column] = converted
            if fail_rate > 0:
                fail_rates[column] = fail_rate
        return df, fail_rates

    def _convert_series(self, series: pd.Series, target_type: str, column_name: Optional[str] = None) -> Tuple[pd.Series, float]:
        cleaned = series.astype(str).str.strip()
        cleaned = cleaned.replace({val: np.nan for val in self.na_values})

        if target_type == "percent":
            percent_mask = cleaned.str.contains("%", na=False)
            cleaned = (
                cleaned.str.replace("%", "", regex=False)
                .str.replace("??", "", regex=False)
                .str.replace(",", "")
            )
            numeric = pd.to_numeric(cleaned, errors="coerce")
            if percent_mask.any():
                numeric.loc[percent_mask] = numeric.loc[percent_mask] / 100.0
            abs_max = float(numeric.abs().max(skipna=True)) if numeric.notna().any() else 0.0
            if not percent_mask.any() and abs_max > 1.5:
                numeric = numeric / 100.0
                abs_max = abs_max / 100.0
            if abs_max > 0 and abs_max <= 0.01 and column_name:
                print(f"[WARN] ? {column_name} ???? {abs_max*100:.2f}% ???????? 100")
        elif target_type == "amount":
            numeric = cleaned.apply(self._parse_amount)
        elif target_type == "date":
            numeric = pd.to_datetime(cleaned, errors="coerce")
        elif target_type == "auto":
            dtype = infer_dtype(cleaned, skipna=True)
            if dtype in ("integer", "floating", "mixed-integer-float"):
                numeric = pd.to_numeric(cleaned, errors="coerce")
            elif dtype == "datetime":
                numeric = pd.to_datetime(cleaned, errors="coerce")
            else:
                numeric = cleaned
        else:
            numeric = pd.to_numeric(cleaned, errors="coerce")

        fail_rate = float(numeric.isna().mean())
        return numeric, fail_rate

    def _parse_amount(self, value: str) -> Optional[float]:
        if not value or value in self.na_values:
            return np.nan
        try:
            for unit, multiplier in self.unit_rules.items():
                if value.endswith(unit):
                    value = value.replace(unit, "")
                    return float(value.replace(",", "")) * multiplier
            return float(value.replace(",", ""))
        except Exception:
            return np.nan


class ExcelParser:
    """组合多引擎加载、列名映射、类型转换的综合解析器。"""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = DEFAULT_PARSE_CONFIG.copy()
        if config:
            self.config.update(config)
        self.loader = MultiEngineExcelLoader(self.config)
        self.column_normalizer = ColumnNormalizer(self.config.get("column_aliases", {}))
        self.value_converter = ValueConverter(self.config)

    def parse_file(self, file_path: str) -> List[Tuple[pd.DataFrame, ParseDiagnostics]]:
        sheets = self.loader.read_file(file_path)
        results: List[Tuple[pd.DataFrame, ParseDiagnostics]] = []
        sheet_policy = self.config.get("sheet_policy", "all")

        target_sheets = list(sheets.keys())
        if isinstance(sheet_policy, list):
            target_sheets = [s for s in sheet_policy if s in sheets]
        elif sheet_policy == "first":
            target_sheets = target_sheets[:1]
        elif sheet_policy == "largest":
            target_sheets = sorted(target_sheets, key=lambda s: len(sheets[s]), reverse=True)[:1]

        for sheet_name in target_sheets:
            df_raw = self._prepare_dataframe(sheets[sheet_name])
            columns_mapped, unmapped = self.column_normalizer.map_columns(df_raw.columns)
            unique_columns, duplicate_counts = self._ensure_unique_columns(columns_mapped)
            df_raw.columns = unique_columns

            df_converted, fail_rates = self.value_converter.convert(df_raw)
            diag = ParseDiagnostics(
                file_path=file_path,
                sheet_name=sheet_name,
                rows=len(df_converted),
                columns=len(df_converted.columns),
                present_columns=df_converted.columns.tolist(),
                unmapped_columns=unmapped,
                conversion_failures=fail_rates,
            )
            if duplicate_counts:
                duplicates_text = ", ".join(
                    f"{name}×{count}"
                    for name, count in duplicate_counts.items()
                )
                diag.notes.append(f"检测到重复列名，已自动重命名: {duplicates_text}")
            results.append((df_converted, diag))
        return results

    def _prepare_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """如果没有显式表头，尝试自动识别有效表头行。"""
        if df is None or df.empty:
            return pd.DataFrame()
        default_columns = all(
            isinstance(col, (int, np.integer)) or str(col).isdigit()
            for col in df.columns
        )
        if not default_columns:
            return df
        header_row = self._detect_header_row(df)
        if header_row is None or header_row >= len(df) - 1:
            return df
        header = df.iloc[header_row].fillna("").astype(str).tolist()
        remaining = df.iloc[header_row + 1 :].copy()
        remaining.columns = header
        remaining = remaining.reset_index(drop=True)
        return remaining

    def _ensure_unique_columns(self, columns: Sequence[str]) -> Tuple[List[str], Dict[str, int]]:
        """
        确保列名唯一，仿照pandas的行为为重复列添加“.1/.2”后缀。
        返回去重后的列名和重复列计数字典。
        """
        counts: Dict[str, int] = {}
        unique: List[str] = []
        for idx, name in enumerate(columns):
            base = name if name not in ("", None) else f"Unnamed_{idx+1}"
            count = counts.get(base, 0)
            unique_name = base if count == 0 else f"{base}.{count}"
            counts[base] = count + 1
            unique.append(unique_name)
        duplicates = {name: count for name, count in counts.items() if count > 1}
        return unique, duplicates

    def _detect_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """依据别名映射在前若干行内寻找最可能的表头。"""
        alias_keys = self.column_normalizer.standard_keys
        best_row = None
        best_score = 0
        max_rows = min(len(df), 30)

        for idx in range(max_rows):
            row = df.iloc[idx]
            cleaned = [ColumnNormalizer._normalize_key(val) for val in row.fillna("").astype(str)]
            score = sum(1 for cell in cleaned if cell in alias_keys)
            if score > best_score:
                best_score = score
                best_row = idx

        if best_score >= 2:
            return best_row

        for idx in range(len(df)):
            row = df.iloc[idx]
            if row.dropna().astype(str).str.strip().any():
                return idx
        return None


def load_excel_sources(paths: Sequence[str], config: Optional[Dict[str, Any]] = None) -> ParsedData:
    parser = ExcelParser(config)
    frames: List[pd.DataFrame] = []
    diagnostics: List[ParseDiagnostics] = []
    unavailable: List[str] = []

    for path in paths:
        if not path:
            continue
        try:
            parsed = parser.parse_file(path)
        except Exception as exc:
            diag = ParseDiagnostics(
                file_path=path,
                sheet_name="N/A",
                rows=0,
                columns=0,
                notes=[f"解析失败: {exc}"],
            )
            diagnostics.append(diag)
            if parser.config.get("strict_mode", False):
                raise
            continue

        for df, diag in parsed:
            frames.append(df)
            diagnostics.append(diag)
            for column, fail_rate in diag.conversion_failures.items():
                threshold = parser.config.get("percent_threshold", 0.8)
                if fail_rate >= threshold:
                    unavailable.append(column)
                    diag.notes.append(
                        f"列 '{column}' 由于转换失败率 {fail_rate:.0%} 被标记为不可用"
                    )

    combined = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    return ParsedData(combined, diagnostics, sorted(set(unavailable)))
