# -*- coding: utf-8 -*-
"""
带参数因子分析报告模块。

该模块承载 ParameterizedFactorAnalyzer 生成 TXT/CSV 报告的逻辑，
主类只需调用公开函数即可，避免大段字符串常量留在主脚本中。
"""

import math
import os
from datetime import datetime
from html import escape

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .fa_config import DEFAULT_DATA_FILE, DEFAULT_DATA_FILES, build_report_path
from .fa_report_utils import (
    HTMLReportBuilder,
    render_metric_cards,
    render_table,
    render_alert,
    render_list,
)

DATA_FILE_LABEL = (
    ", ".join(DEFAULT_DATA_FILES)
    if DEFAULT_DATA_FILES
    else (DEFAULT_DATA_FILE or "（未配置数据文件）")
)


def _fmt_percent(value, decimals=1):
    if pd.isna(value):
        return "--"
    return f"{value:.{decimals}%}"


def _fmt_float(value, decimals=3):
    if pd.isna(value):
        return "--"
    return f"{value:.{decimals}f}"


def _fmt_score(value):
    if pd.isna(value):
        return "--"
    return f"{value:.1f}"


def _fa_generate_parameterized_report(self):
    """生成带参数因子综合分析报告并输出所有相关文件。"""
    print("开始生成带参数因子综合分析报告...")

    factor_results = {}
    for factor in self.factor_list:
        print(f"分析因子: {factor}")
        results = self.calculate_comprehensive_metrics(factor)
        if results:
            factor_results[factor] = results

    if not factor_results:
        print("错误: 没有有效的带参数因子分析结果")
        return None

    scores_df = self.score_factors(factor_results)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_filename = f'带参数因子综合分析报告_{timestamp}.html'
    report_path = build_report_path(report_filename)

    builder = HTMLReportBuilder("带参数因子综合分析详细报告")
    builder.set_meta(
        [
            ("生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("数据文件", DATA_FILE_LABEL),
            ("核心指标", "最大回撤 / 年化收益 / 夏普 / 波动"),
        ]
    )

    positive_factors = scores_df[scores_df['因子方向'] == '正向']
    negative_factors = scores_df[scores_df['因子方向'] == '负向']

    overview_cards = render_metric_cards(
        [
            ("总因子数量", str(len(self.factor_list))),
            ("有效分析因子", str(len(factor_results))),
            ("平均年化收益率", _fmt_float(scores_df['年化收益率'].mean(), 3)),
            ("平均年化夏普比率", _fmt_float(scores_df['年化夏普比率'].mean(), 3)),
            ("平均最大回撤", _fmt_percent(scores_df['最大回撤'].mean(), 1)),
        ]
    )
    builder.add_section("整体概览", overview_cards)

    ranking_columns = [
        '排名',
        '因子名称',
        '参数区间',
        '综合得分',
        '相邻平滑后得分',
        '年化收益率',
        '年化夏普比率',
        '平均每笔收益率',
        '最大回撤',
        '交易日数量',
        '样本数量',
        '数据年份',
    ]
    ranking_headers = [
        '排名',
        '因子',
        '参数区间',
        '综合得分',
        '相邻平滑后得分',
        '年化收益率',
        '年化夏普',
        '平均每笔交易收益率',
        '最大回撤',
        '交易日数量',
        '样本数量',
        '数据年份',
    ]
    ranking_formatters = {
        '排名': lambda x: str(int(x)),
        '综合得分': _fmt_score,
        '相邻平滑后得分': _fmt_score,
        '年化收益率': lambda x: _fmt_percent(x, 2),
        '年化夏普比率': _fmt_float,
        '平均每笔收益率': lambda x: _fmt_percent(x, 2),
        '最大回撤': lambda x: _fmt_percent(x, 1),
        '交易日数量': lambda x: str(int(x)) if pd.notna(x) else "--",
        '样本数量': lambda x: str(int(x)) if pd.notna(x) else "--",
        '数据年份': lambda x: str(x) if pd.notna(x) else "--",
    }

    def _wrap_subcard(title, description, inner_html):
        desc_html = f"<p class='muted'>{escape(description)}</p>" if description else ""
        return f"<div class='sub-card'><h3>{escape(title)}</h3>{desc_html}{inner_html}</div>"

    def _build_ranking_block(df, empty_text, description, apply_positive_highlight=False):
        if df.empty:
            return render_alert(empty_text, level="warn")
        ordered_df = df.sort_values('综合得分', ascending=False).reset_index().rename(columns={'index': '__orig_index'})
        ranking_df = ordered_df[
            [
                '因子名称',
                '参数区间',
                '综合得分',
                '相邻平滑后得分',
                '年化收益率',
                '年化夏普比率',
                '平均每笔收益率',
                '最大回撤',
                '交易日数量',
                '样本数量',
                '数据年份',
            ]
        ].copy()
        ranking_df.insert(0, '排名', range(1, len(ranking_df) + 1))
        cell_classes = {}
        if apply_positive_highlight:
            smoothed_rank = df['相邻平滑后得分'].rank(method='min', ascending=False)
            for pos, row in ordered_df.iterrows():
                if pos >= 10:
                    continue
                orig_idx = row['__orig_index']
                if smoothed_rank.loc[orig_idx] > 10:
                    cell_classes.setdefault(pos, {})['相邻平滑后得分'] = 'highlight-alert'
            metric_cols = ['年化收益率', '年化夏普比率', '平均每笔收益率', '最大回撤']
            for col in metric_cols:
                values = pd.to_numeric(ordered_df[col], errors='coerce')
                if col == '最大回撤':
                    top_positions = values.nsmallest(3).index
                else:
                    top_positions = values.nlargest(3).index
                for pos_idx in top_positions:
                    cell_classes.setdefault(pos_idx, {})[col] = 'highlight-blue'
        for pos in range(min(10, len(ordered_df))):
            cell_classes.setdefault(pos, {})['综合得分'] = 'score-emphasis'
            cell_classes.setdefault(pos, {})['因子名称'] = 'factor-emphasis'
        table_html = render_table(
            ranking_df,
            ranking_columns,
            headers=ranking_headers,
            formatters=ranking_formatters,
            cell_classes=cell_classes or None,
        )
        best_row = ordered_df.iloc[0]
        highlight = render_alert(
            f"最佳区间：{escape(best_row['因子名称'])} {escape(best_row['参数区间'])}，"
            f"相邻平滑后得分 { _fmt_score(best_row['相邻平滑后得分']) }。",
        level="info",
    )
        return _wrap_subcard(description, "根据综合得分与收益/风险指标排序", table_html + highlight)

    leaderboard_html = (
        _build_ranking_block(positive_factors, "暂无正向因子满足条件", "正向参数区间排行榜", apply_positive_highlight=True)
        + _build_ranking_block(negative_factors, "暂无负向因子满足条件", "负向参数区间排行榜（反向使用）")
    )
    builder.add_section("参数区间排行榜", leaderboard_html)

    def _build_top_list(frame, badge_class):
        if frame.empty:
            return render_alert("暂无可推荐的参数区间。", level="warn")
        items_html = []
        for i, (_, factor) in enumerate(frame.iterrows(), 1):
            stats = [
                f"年化收益率：{_fmt_percent(factor['年化收益率'], 2)}",
                f"年化夏普比率：{_fmt_float(factor['年化夏普比率'], 3)}",
                f"年化收益标准差：{_fmt_float(factor['年化收益标准差'], 3)}",
                f"平均每笔收益率：{_fmt_percent(factor.get('平均每笔收益率'), 2)}",
                f"最大回撤：{_fmt_percent(factor['最大回撤'], 1)}",
                f"交易日数量：{int(factor['交易日数量']) if pd.notna(factor['交易日数量']) else '--'}",
                f"样本数量：{int(factor['样本数量']) if pd.notna(factor['样本数量']) else '--'}",
            ]
            items_html.append(
                f"<li><strong>第{i}名：{escape(factor['因子名称'])} | {escape(factor['参数区间'])}</strong>"
                f"<span class='badge {badge_class}'>相邻平滑后得分 {_fmt_score(factor['相邻平滑后得分'])}</span>"
                f"{render_list(stats)}</li>"
            )
        return f"<ul class='top-list'>{''.join(items_html)}</ul>"

    builder.add_section("正向区间 TOP5", _wrap_subcard("正向精选", "重点关注收益与稳健兼顾的区间", _build_top_list(positive_factors.head(5), "badge-positive")))
    builder.add_section("负向区间 TOP5（对冲/反向）", _wrap_subcard("负向精选", "用于风险对冲或反向策略", _build_top_list(negative_factors.head(5), "badge-negative")))

    detail_cards = []
    all_factors_sorted = scores_df.sort_values('相邻平滑后得分', ascending=False)
    for _, factor_row in all_factors_sorted.iterrows():
        if factor_row['相邻平滑后得分'] >= 9:
            rating = "A级（优秀）"
        elif factor_row['相邻平滑后得分'] >= 8:
            rating = "B+级（良好）"
        elif factor_row['相邻平滑后得分'] >= 6:
            rating = "B级（一般）"
        else:
            rating = "C级（较差）"
        direction_tag = "tag-positive" if factor_row['因子方向'] == '正向' else "tag-negative"
        metric_items = [
            f"年化收益率：{_fmt_percent(factor_row['年化收益率'], 1)}",
            f"年化收益标准差：{_fmt_float(factor_row['年化收益标准差'], 3)}",
            f"年化夏普比率：{_fmt_float(factor_row['年化夏普比率'], 3)}",
            f"平均每笔收益率：{_fmt_percent(factor_row.get('平均每笔收益率'), 1)}",
            f"最大回撤：{_fmt_percent(factor_row['最大回撤'], 1)}",
            f"交易日数量：{int(factor_row['交易日数量']) if pd.notna(factor_row.get('交易日数量')) else '--'}",
            f"样本数量：{int(factor_row['样本数量']) if pd.notna(factor_row.get('样本数量')) else '--'}",
        ]
        group_html = ""
        factor_name = factor_row['因子名称']
        if factor_name in factor_results:
            group_stats = factor_results[factor_name]['group_stats']
            param_range = factor_row['参数区间']
            group_data = group_stats[group_stats['参数区间'] == param_range]
            if len(group_data) > 0:
                group = group_data.iloc[0]
                group_items = [
                    f"平均收益：{_fmt_percent(group['平均收益'], 1)}",
                    f"收益标准差：{_fmt_float(group['收益标准差'], 3)}",
                    f"最大回撤：{_fmt_percent(group['最大回撤'], 1)}",
                    f"年化收益率：{_fmt_percent(group['年化收益率'], 1)}",
                    f"年化收益标准差：{_fmt_float(group['年化收益标准差'], 3)}",
                    f"年化夏普比率：{_fmt_float(group['年化夏普比率'], 3)}",
                ]
                if '年化索提诺比率' in group and pd.notna(group['年化索提诺比率']):
                    group_items.append(f"年化索提诺比率：{_fmt_float(group['年化索提诺比率'], 3)}")
                if '日度收益均值' in group and pd.notna(group['日度收益均值']):
                    group_items.append(f"日度收益均值：{_fmt_percent(group['日度收益均值'], 1)}")
                if '日度收益波动' in group and pd.notna(group['日度收益波动']):
                    group_items.append(f"日度收益波动：{_fmt_float(group['日度收益波动'], 3)}")
                if '交易日数量' in group and pd.notna(group['交易日数量']):
                    trade_days_val = int(group['交易日数量'])
                    group_items.append(f"交易日数量：{trade_days_val}")
                if '样本数量' in group and pd.notna(group['样本数量']):
                    group_items.append(f"样本数量：{int(group['样本数量'])}")
                if '观测区间' in group and isinstance(group['观测区间'], str) and group['观测区间']:
                    group_items.append(f"观察区间：{group['观测区间']}")
                if '观测期年数' in group and pd.notna(group['观测期年数']):
                    group_items.append(f"观测期（年）：{_fmt_float(group['观测期年数'], 3)}")
                if '年化收益估算方式' in group and isinstance(group['年化收益估算方式'], str):
                    group_items.append(f"年化估算方式：{group['年化收益估算方式']}")
                sample_hint = group.get('年化样本提示')
                if isinstance(sample_hint, str) and sample_hint.strip():
                    group_items.append(f"样本提示：{sample_hint.strip()}")
                group_html = f"<div class='sub-block'><h4>分组详细数据</h4>{render_list(group_items)}</div>"

        detail_cards.append(
            f"<div class='factor-card'>"
            f"<h3>{escape(factor_row['因子名称'])} | {escape(factor_row['参数区间'])}</h3>"
            f"<p><span class='tag {direction_tag}'>{escape(factor_row['因子方向'])}</span>"
            f"<span class='tag'>{escape(rating)}</span>"
            f"<span class='tag'>相邻平滑后得分 {_fmt_score(factor_row['相邻平滑后得分'])}</span>"
            f"<span class='tag muted'>综合得分 {_fmt_score(factor_row['综合得分'])}</span></p>"
            f"{render_list(metric_items)}"
            f"{group_html}"
            "</div>"
        )

    builder.add_section(
        "详细参数区间分析",
        "".join(detail_cards) if detail_cards else render_alert("暂无可用的因子分析详情。", level="warn"),
    )

    strategy_items = [
        "重点跟踪排名前列的正向区间，可按相邻平滑后得分分配 20%~40% 的组合权重。",
        "若需要对冲，可挑选负向区间作为保护腿，控制在 10% 左右的权重。",
        "不同参数区间之间保持风格多元，避免集中在单一市值或板块。",
        "实时监控最大回撤和收益波动，若连续恶化需及时调权或剔除。",
        "固定周期（如每月/每季度）重新评估参数区间有效性与稳健性。",
        "将高分区间与基本面或风控因子结合使用，提高策略抗压能力。",
    ]
    builder.add_section("投资策略建议", render_list(strategy_items))

    risk_items = [
        "历史表现不代表未来收益，需结合实时行情二次验证。",
        "带参数因子的有效性对数据质量与样本代表性高度敏感。",
        "参数区间若长期失效应及时回测与下调权重，避免情绪化交易。",
        "建议与非参数分析、行业洞察等多维信号交叉验证。",
        "分散参数区间，控制单一策略的资金暴露，提高组合韧性。",
        "持续监控成交与流动性变化，必要时暂停使用异常区间。",
    ]
    builder.add_section("风险提示", render_list(risk_items))

    html_content = builder.render()
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"带参数因子综合分析报告已保存到 '{report_path}'")

    data_filename = f'带参数因子分析数据_{timestamp}.xlsx'
    data_path = build_report_path(data_filename)
    scores_df.to_excel(data_path, index=False)
    print(f"详细数据已保存到 '{data_path}'")

    try:
        _apply_param_csv_highlight(data_path, scores_df, already_excel=True)
    except Exception as e:
        print(f"[WARN] 生成高亮Excel失败: {e}")

    return report_path


def _apply_param_csv_highlight(file_path, df, already_excel=False):
    """
    生成一个Excel副本，并对关键列进行分位高亮
    """
    if df is None or df.empty:
        return

    if already_excel:
        xlsx_path = file_path
    else:
        xlsx_path = os.path.splitext(file_path)[0] + '.xlsx'
        df.to_excel(xlsx_path, index=False)

    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_to_col = {cell.value: cell.column for cell in ws[1]}
        target_cols = ['年化收益率', '综合得分', '相邻平滑后得分']

        blue_fill = PatternFill(start_color='FF4F81BD', end_color='FF4F81BD', fill_type='solid')
        red_fill = PatternFill(start_color='FFFF6B6B', end_color='FFFF6B6B', fill_type='solid')

        for col_name in target_cols:
            col_idx = header_to_col.get(col_name)
            if not col_idx:
                continue
            values = []
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    val = float(cell.value)
                except (TypeError, ValueError):
                    continue
                values.append((row, val))
            if not values:
                continue
            highlight_count = max(1, math.ceil(len(values) * 0.2))
            sorted_desc = sorted(values, key=lambda x: x[1], reverse=True)
            sorted_asc = sorted(values, key=lambda x: x[1])
            top_rows = {row for row, _ in sorted_desc[:highlight_count]}
            bottom_rows = {row for row, _ in sorted_asc[:highlight_count]}

            for row, _ in values:
                cell = ws.cell(row=row, column=col_idx)
                if row in top_rows:
                    cell.fill = blue_fill
                elif row in bottom_rows:
                    cell.fill = red_fill
                else:
                    cell.fill = PatternFill(fill_type=None)

        wb.save(xlsx_path)
        print(f"[INFO] 已生成带高亮的Excel: {xlsx_path}")
    except Exception as e:
        print(f"[WARN] Excel高亮处理失败: {e}")
